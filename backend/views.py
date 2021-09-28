import os
from django.conf import settings
from django.core.exceptions import PermissionDenied
import locale
from typing import ContextManager
from django.forms.models import ALL_FIELDS
from django.http.response import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font, numbers
import pytz
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required , user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls.base import reverse_lazy
from accounts.models import CustomUser
from datetime import datetime
from calendar import calendar, monthrange
from django.views.generic import(
    ListView,
    CreateView
)
from .models import ButirCKP, DokumenCKP, MasterKegiatan, MasterButirKegiatan
from .forms import ButirCKPCreationFormR, DokumenCKPCreationForm, KegiatanCreationForm, AdminKegiatanCreationForm, PenilaianButirCKPForm

# <------------------------------------------ Permission and Authorization ------------------------------->

def edit_check(user, dokumen_ckp):
    if ((dokumen_ckp.pegawai != user) and (user.jabatan_kantor not in CustomUser.KASI_LIST)):
        return True

def error_403(request, exception):
    return render(request, 'backend/tidakdiizinkan.html')

def tidak_diizinkan(request):
    return render(request, 'backend/tidakdiizinkan.html')

# <------------------------------------------ SICAKEP - Kegiatan ----------------------------------------->


@login_required
def home(request):
    context = {
        'masterkegiatans': MasterKegiatan.objects.all(),
        'title': 'Master Kegiatan',
    }
    return render(request, 'backend/kegiatan/masterkegiatan.html', context)
    # return render(request, 'backend/index.html')


# function-based view
@login_required
def masterkegiatan(request):
    context = {
        'masterkegiatans': MasterKegiatan.objects.all(),
        'title': 'Master Kegiatan',
    }
    return render(request, 'backend/kegiatan/masterkegiatan.html', context)


class MasterKegiatanListView(LoginRequiredMixin, ListView):
    model = MasterKegiatan
    template_name = 'backend/kegiatan/masterkegiatan.html'
    context_object_name = 'masterkegiatans'


@login_required
def tambah_kegiatan(request):
    if request.method == 'POST' and request.user.is_superuser:
        form = AdminKegiatanCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(
                request, f'SUKSES! Kegiatan baru berhasil ditambahkan, Sekarang anda dapat menggunakannya di CKP')
            if 'add_another' in request.POST:
                return redirect('sicakep-tambahkegiatan')
            else:
                return redirect('sicakep-masterkegiatan')
    elif request.method == 'POST' and not request.user.is_superuser:
        form = KegiatanCreationForm(request.POST)
        if form.is_valid():
            fs = form.save(commit=False)
            fs.author = request.user
            fs.subject_matter = request.user.jabatan_kantor
            fs.save()
            messages.success(
                request, f'SUKSES! Kegiatan baru berhasil ditambahkan, Sekarang anda dapat menggunakannya di CKP')
            if 'add_another' in request.POST:
                return redirect('sicakep-tambahkegiatan')
            else:
                return redirect('sicakep-masterkegiatan')
    else:
        form = KegiatanCreationForm()

    jabatan_kantor_choices = CustomUser.get_jabatan_kantor_dropdown()
    author_list = CustomUser.objects.all()
    butir_kegiatan_list = MasterButirKegiatan.objects.all()

    context = {
        'jabatan_kantor_choices': jabatan_kantor_choices,
        'form': form,
        'author_list': author_list,
        'butir_kegiatan_list': butir_kegiatan_list,
        'title': 'Tambah Kegiatan',
    }
    return render(request, 'backend/kegiatan/tambahkegiatan.html', context)


class TambahKegiatanView(LoginRequiredMixin, CreateView):
    model = MasterKegiatan
    template_name = 'backend/kegiatan/tambahkegiatan.html'
    fields = ['uraian_kegiatan', 'subject_matter', 'butir_kegiatan']
    success_url = reverse_lazy('sicakep-masterkegiatan')

    def get_context_data(self, **kwargs):
        context = super(TambahKegiatanView, self).get_context_data(**kwargs)
        context['jabatan_kantor_choices'] = CustomUser.get_jabatan_kantor_dropdown()
        return context

    def form_valid(self, form):
        form.instance.author = self.request.user
        return super().form_valid(form)


@login_required
def detail_kegiatan(request, pk):
    kegiatan = MasterKegiatan.objects.get(id=pk)
    jabatan_kantor_choices = CustomUser.get_jabatan_kantor_dropdown()

    context = {
        'jabatan_kantor_choices': jabatan_kantor_choices,
        'kegiatan': kegiatan,
        'title': 'Detail Kegiatan',
    }
    return render(request, 'backend/kegiatan/detailkegiatan.html', context)


@login_required
def update_kegiatan(request, pk):
    kegiatan = MasterKegiatan.objects.get(id=pk)
    if request.user.is_superuser:
        form = AdminKegiatanCreationForm(instance=kegiatan)
    else:
        form = KegiatanCreationForm(instance=kegiatan)
    jabatan_kantor_choices = CustomUser.get_jabatan_kantor_dropdown()
    author_list = CustomUser.objects.all()
    butir_kegiatan_list = MasterButirKegiatan.objects.all()

    if request.method == 'POST' and request.user.is_superuser:
        form = AdminKegiatanCreationForm(request.POST, instance=kegiatan)
        if form.is_valid:
            form.save()
            messages.success(
                request, f'SUKSES! Kegiatan berhasil diupdate')
            return redirect('sicakep-masterkegiatan')

    if request.method == 'POST' and not request.user.is_superuser:
        form = KegiatanCreationForm(request.POST, instance=kegiatan)
        if form.is_valid:
            fs = form.save(commit=False)
            fs.author = request.user
            fs.subject_matter = request.user.jabatan_kantor
            fs.save()
            messages.success(
                request, f'SUKSES! Kegiatan berhasil diupdate')
            return redirect('sicakep-masterkegiatan')

    context = {
        'jabatan_kantor_choices': jabatan_kantor_choices,
        'kegiatan': kegiatan,
        'form': form,
        'author_list': author_list,
        'butir_kegiatan_list': butir_kegiatan_list,
        'title': 'Update Kegiatan',
    }
    return render(request, 'backend/kegiatan/updatekegiatan.html', context)


@login_required
def delete_kegiatan(request, pk):
    kegiatan = MasterKegiatan.objects.get(id=pk)
    if request.method == 'POST':
        if 'hapus' in request.POST:
            kegiatan.delete()
            return redirect('sicakep-masterkegiatan')

    return redirect('sicakep-masterkegiatan')

# <------------------------------------------ END SICAKEP - Kegiatan ----------------------------------------->

# <------------------------------------------ SICAKEP - CKP ----------------------------------------->


@login_required
def daftar_ckp(request):

    if request.user.jabatan_kantor not in CustomUser.KASI_LIST:
        ckp_list = DokumenCKP.objects.filter(pegawai=request.user)
    else:
        ckp_list = DokumenCKP.objects.all()

    context = {
        'ckp_list': ckp_list,
        'title': 'Daftar CKP Pegawai',
    }

    return render(request, 'backend/ckp/daftarckp.html', context)


@login_required
def tambah_ckp_pegawai(request):
    if request.method == 'POST':
        form = DokumenCKPCreationForm(request.POST)
        if form.is_valid:
            fs = form.save(commit=False)
            # fs.periode = datetime(datetime.now().year, int(request.POST.get(
            #     'bulan', False)), 1, tzinfo=pytz.timezone("Asia/Jakarta"))
            fs.periode = datetime(datetime.now().year, int(request.POST.get(
                'bulan', False)), 1, hour=12)
            if DokumenCKP.objects.filter(periode=fs.periode, pegawai=form.cleaned_data['pegawai']).exists():
                messages.warning(
                    request, f'Peringatan! CKP ' + form.cleaned_data['pegawai'].get_full_name() + ' untuk Bulan ' + fs.periode.strftime("%B") + ' sudah dibuat, Silahkan tambahkan kegiatan di menu daftar CKP')
                return redirect('sicakep-daftarckp')
            else:    
                fs.save()
                messages.success(
                    request, f'SUKSES! CKP ' + form.cleaned_data['pegawai'].get_full_name() + ' untuk Bulan ' + fs.periode.strftime("%B") + ' berhasil dibuat, Silahkan tambahkan kegiatan di menu daftar CKP')
                return redirect('sicakep-daftarckp')
    else:
        form = DokumenCKPCreationForm()

    bulan_max_list = DokumenCKP.get_daftar_bulan_max()

    if request.user.jabatan_kantor == '1' or request.user.is_superuser:
        pegawai_list = CustomUser.objects.all().exclude(
            is_superuser=True).order_by('first_name')
    elif request.user.jabatan_kantor in CustomUser.KASI_LIST:
        pegawai_list = CustomUser.objects.all().exclude(
            is_superuser=True).exclude(jabatan_kantor='1').order_by('first_name')
    else:
        pegawai_list = CustomUser.objects.filter(id__exact=request.user.id)

    context = {
        'bulan_max_list': bulan_max_list,
        'pegawai_list': pegawai_list,
        'title': 'Tambah CKP Pegawai',
    }

    return render(request, 'backend/ckp/tambahckppegawai.html', context)


@login_required
def detail_ckp(request, pk):
    dokumen_ckp = DokumenCKP.objects.get(id=pk)
    butir_ckp_set = dokumen_ckp.butirckp_set.all()

    context = {
        'dokumen_ckp': dokumen_ckp,
        'butir_ckp_set': butir_ckp_set,
        'title': 'Detail CKP',
    }

    if edit_check(request.user, dokumen_ckp):
        # return redirect('tidak-diizinkan')
        raise PermissionDenied 
    else:
        return render(request, 'backend/ckp/detailckp.html', context)


@login_required
def tambah_kegiatan_ckp(request, pk):
    if request.method == 'POST' and not request.user.is_superuser:
        form = ButirCKPCreationFormR(request.POST)
        if form.is_valid():
            fs = form.save(commit=False)
            fs.kegiatan = MasterKegiatan.objects.get(
                id=request.POST.get('kegiatan', False))
            fs.save()
            fs.dokumen_ckp.add(DokumenCKP.objects.get(id=pk))
            messages.success(
                request, f'SUKSES! Kegiatan baru berhasil ditambahkan')
            if 'add_another' in request.POST:
                return redirect('sicakep-tambahkegiatanckp', pk=pk)
            else:
                return redirect('sicakep-detailckp', pk=pk)
    else:
        form = ButirCKPCreationFormR()

    dokumen_ckp = DokumenCKP.objects.get(id=pk)
    kegiatan_list = MasterKegiatan.objects.all()

    context = {
        'dokumen_ckp': dokumen_ckp,
        'kegiatan_list': kegiatan_list,
        'title': 'Tambah Butir Kegiatan CKP',
    }
    return render(request, 'backend/ckp/tambahkegiatanckp.html', context)

@login_required
def delete_dokumen_ckp(request, pk):
    dokumen_ckp = DokumenCKP.objects.get(id=pk)
    if request.method == 'POST':
        if 'hapus' in request.POST:
            dokumen_ckp.delete()
            return redirect('sicakep-daftarckp')

    return redirect('sicakep-daftarckp')

@login_required
def delete_butir_ckp(request, btpk, dkpk):
    butir_ckp = ButirCKP.objects.get(id=btpk)
    if request.method == 'POST':
        if 'hapus' in request.POST:
            butir_ckp.delete()
            return redirect('sicakep-detailckp', dkpk)

    return redirect('sicakep-detailckp', dkpk)

@login_required
def penilaian_ckp(request, pk):
    butir_ckp = ButirCKP.objects.get(id=pk)
    dokumen_ckp = butir_ckp.dokumen_ckp.first()

    if request.method == 'POST':
        form = PenilaianButirCKPForm(request.POST, instance=butir_ckp)
        if form.is_valid():
            fs = form.save(commit=False)
            fs.jenis_butir_ckp = butir_ckp.jenis_butir_ckp
            fs.kegiatan = butir_ckp.kegiatan
            fs.target = butir_ckp.target
            fs.save()
            messages.success(request, f'SUKSES! Penilaian berhasil dilakukan')
            return redirect('sicakep-detailckp', dokumen_ckp.id)
        else:
            print('form tidak valid')
    else:
        form = PenilaianButirCKPForm()

    context = {
        'butir_ckp': butir_ckp,
        'dokumen_ckp': dokumen_ckp,
        'title': 'Penilaian CKP',
    }

    if edit_check(request.user, dokumen_ckp):
        raise PermissionDenied
    else:
        return render(request, 'backend/ckp/entripenilaian.html', context)


@login_required
def export_xls_ckp(request, pk):
    dokumen_ckp = DokumenCKP.objects.get(id=pk)
    butir_ckp_set = dokumen_ckp.butirckp_set.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    locale.setlocale(locale.LC_TIME, "IND")
    file_name = 'CKP ' + dokumen_ckp.pegawai.get_full_name() + ' ' + \
        dokumen_ckp.periode.strftime("%B %Y") + '.xlsx'
    response['Content-Disposition'] = 'attachment; filename="' + file_name + '"'
    # wb = load_workbook(os.path.join(settings.BASE_DIR,
    #                    'media/excel/TEMPLATE CKP.xlsx'))

    wb = Workbook()

    last_day_of_month = monthrange(
        dokumen_ckp.periode.year, dokumen_ckp.periode.month)[1]

    # Sheet CKP-T -----------------------------------------------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "CKP-T"
    # ws1 = wb["CKP-T"]

    # Inisialisai style
    default_style = NamedStyle(name="default_style")
    default_style.font = Font(name='Arial')
    default_style.alignment = Alignment(vertical='center')

    bd = Border(
        bottom=Side(
            border_style='thin',
            color='000000'
        ),
        top=Side(
            border_style='thin',
            color='000000'
        ),
        left=Side(
            border_style='thin',
            color='000000'
        ),
        right=Side(
            border_style='thin',
            color='000000'
        ),
        outline=Side(
            border_style='thin',
            color='000000'
        ),
    )

    tengahboldborder = NamedStyle(name="tengahboldborder")
    tengahboldborder.font = Font(bold=True, name="Arial")
    tengahboldborder.alignment = Alignment(horizontal="center", vertical="center")
    tengahboldborder.border = bd

    tengahbold = NamedStyle(name="tengabold")
    tengahbold.font = Font(bold=True, name='Arial')
    tengahbold.alignment = Alignment(horizontal="center", vertical="center")

    header1 = NamedStyle(name="header1")
    header1.font = Font(bold=True, name='Arial')
    header1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header1.fill = PatternFill(fill_type="solid", start_color='cccccc')
    header1.border = bd

    header2 = NamedStyle(name="header2")
    header2.font = Font(name="Arial")
    header2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header2.fill = PatternFill(fill_type="solid", start_color='cccccc')
    header2.border = bd

    header3 = NamedStyle(name="header3")
    header3.font = Font(underline='single', bold=True, name='Arial')
    header3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header3.border = bd
    
    isitabelkiri = NamedStyle("isitabelkiri")
    isitabelkiri.font = Font(name="Arial")
    isitabelkiri.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    isitabelkiri.border = bd

    isitabelkanan = NamedStyle("isitabelkanan")
    isitabelkanan.font = Font(name="Arial")
    isitabelkanan.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    isitabelkanan.border = bd

    isitabeltengah = NamedStyle("isitabetengah")
    isitabeltengah.font = Font(name="Arial")
    isitabeltengah.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    isitabeltengah.border = bd

    boldkiri = NamedStyle("boldkiri")
    boldkiri.font = Font(bold=True, name="Arial")
    boldkiri.alignment = Alignment(horizontal="left", wrap_text=True)

    boldtengah = NamedStyle("boldtengah")
    boldtengah.font = Font(bold=True, name="Arial")
    boldtengah.alignment = Alignment(horizontal="center", wrap_text=True)

    tengah = NamedStyle("tengah")
    tengah.font = Font(name="Arial")
    tengah.alignment = Alignment(horizontal="center", wrap_text=True)


    ws1.sheet_format.defaultRowHeight = 20
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 17
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 11
    ws1.column_dimensions['H'].width = 50

    # for cells in ws1.rows:
    #     for cell in cells:
    #         cell.style = default_style

    # Merge Cell
    ws1.merge_cells('A2:H2')
    ws1.merge_cells('A4:B4')
    ws1.merge_cells('A5:B5')
    ws1.merge_cells('A6:B6')
    ws1.merge_cells('A7:B7')
    ws1.merge_cells('B9:C10')
    ws1.merge_cells('A9:A10')
    ws1.merge_cells('D9:D10')
    ws1.merge_cells('E9:E10')
    ws1.merge_cells('F9:F10')
    ws1.merge_cells('G9:G10')
    ws1.merge_cells('H9:H10')
    ws1.merge_cells('B11:C11')
    ws1.merge_cells('A12:H12')

    for col in ws1['A4':'C7']:
        for cell in col:
            cell.style = default_style

    ws1.row_dimensions[1].height = 30
    ws1['H1'].style = tengahboldborder
    ws1.row_dimensions[2].height = 25
    ws1['A2'].style = tengahbold
    ws1.row_dimensions[9].height = 25
    ws1.row_dimensions[10].height = 25
    for col in ws1['A9':'H10']:
        for cell in col:
            cell.style = header1
    ws1.row_dimensions[11].height = 20
    for col in ws1['A11':'H11']:
        for cell in col:
            cell.style = header2
    ws1.row_dimensions[12].height = 20
    for col in ws1['A12':'H12']:
        for cell in col:
            cell.style = header3
    for i in range(4,9):
        ws1.row_dimensions[i].height = 17
    
    ws1['H1'] = "CKP-T"
    ws1['A2'] = "TARGET KINERJA PEGAWAI TAHUN " + str(dokumen_ckp.periode.year)
    ws1['A4'] = "Satuan Organisasi"
    ws1['C4'] = ": BPS Kabupaten Kuantan Singingi"
    ws1['A5'] = "Nama"
    ws1['C5'] = ": " + dokumen_ckp.pegawai.get_full_name()
    ws1['A6'] = "Jabatan"
    ws1['C6'] = ": " + dokumen_ckp.pegawai.get_jabatan_kantor_display()
    ws1['A7'] = "Periode"
    ws1['C7'] = ": 1 " + dokumen_ckp.periode.strftime("%B ") + "- " + str(last_day_of_month) + " " +dokumen_ckp.periode.strftime("%B %Y")

    ws1['A9'] = "No"
    ws1['B9'] = "Uraian Kegiatan"
    ws1['D9'] = "Satuan"
    ws1['E9'] = "Target Kuantitas"
    ws1['F9'] = "Kode Butir Kegiatan"
    ws1['G9'] = "Angka Kredit"
    ws1['H9'] = "Keterangan"
    ws1['A11'] = "(1)"
    ws1['B11'] = "(2)"
    ws1['D11'] = "(3)"
    ws1['E11'] = "(4)"
    ws1['F11'] = "(5)"
    ws1['G11'] = "(6)"
    ws1['H11'] = "(7)"
    ws1['A12'] = "UTAMA"

    global baris
    global nomor
    baris = 13
    nomor = 1

    for idx, butir_ckp in enumerate(butir_ckp_set):

        if butir_ckp.jenis_butir_ckp == 'TAMBAHAN':
            continue

        ws1.merge_cells('B'+str(baris)+':C'+str(baris))

        ws1['A'+str(baris)] = nomor
        ws1['B'+str(baris)] = butir_ckp.kegiatan.uraian_kegiatan
        ws1['D'+str(baris)] = butir_ckp.kegiatan.satuan
        ws1['E'+str(baris)] = butir_ckp.target
        if butir_ckp.kegiatan.butir_kegiatan is not None:
            ws1['F'+str(baris)] = butir_ckp.kegiatan.butir_kegiatan.__str__()
            ws1['G'+str(baris)] = "=" + str(butir_ckp.kegiatan.butir_kegiatan.angka_kredit) + "*E"+str(baris)
        ws1['H'+str(baris)] = butir_ckp.keterangan

        for col in ws1['A'+str(baris):'H'+str(baris)]:
            for cell in col:
                cell.style = isitabeltengah

        ws1['B'+str(baris)].style = isitabelkiri
        ws1['E'+str(baris)].style = isitabelkanan
        # ws1['G'+str(baris)].style = isitabelkanan
        ws1['H'+str(baris)].style = isitabelkiri
        
        ws1.row_dimensions[baris].height = 25
        if len(butir_ckp.kegiatan.uraian_kegiatan)>45 :
            ws1.row_dimensions[baris].height = 50
        
        ws1['G'+str(baris)].number_format = '0.000'

        baris +=1
        nomor +=1

    ws1.merge_cells('A'+str(baris)+':H'+str(baris))
    for col in ws1['A'+str(baris):'H'+str(baris)]:
        for cell in col:
            cell.style = header3
    ws1.row_dimensions[baris].height = 20
    baris +=1
    ws1.row_dimensions[baris].height = 20
    for col in ws1['A'+str(baris):'H'+str(baris)]:
        for cell in col:
            cell.style = header3
    ws1['A'+str(baris)] = 'TAMBAHAN'
    ws1.merge_cells('A'+str(baris)+':H'+str(baris))  
    baris +=1
    
    nomor = 1

    for idx, butir_ckp in enumerate(butir_ckp_set):

        if butir_ckp.jenis_butir_ckp == 'UTAMA':
            continue

        ws1.merge_cells('B'+str(baris)+':C'+str(baris))

        ws1['A'+str(baris)] = nomor
        ws1['B'+str(baris)] = butir_ckp.kegiatan.uraian_kegiatan
        ws1['D'+str(baris)] = butir_ckp.kegiatan.satuan
        ws1['E'+str(baris)] = butir_ckp.target
        if butir_ckp.kegiatan.butir_kegiatan is not None:
            ws1['F'+str(baris)] = butir_ckp.kegiatan.butir_kegiatan.__str__()
            ws1['G'+str(baris)] = "=" + str(butir_ckp.kegiatan.butir_kegiatan.angka_kredit) + "*E"+str(baris)
        ws1['H'+str(baris)] = butir_ckp.keterangan

        for col in ws1['A'+str(baris):'H'+str(baris)]:
            for cell in col:
                cell.style = isitabeltengah

        ws1['B'+str(baris)].style = isitabelkiri
        ws1['E'+str(baris)].style = isitabelkanan
        # ws1['G'+str(baris)].style = isitabelkanan
        ws1['H'+str(baris)].style = isitabelkiri
        
        ws1.row_dimensions[baris].height = 25
        if len(butir_ckp.kegiatan.uraian_kegiatan)>45 :
            ws1.row_dimensions[baris].height = 50
        
        ws1['G'+str(baris)].number_format = '0.000'

        baris +=1
        nomor +=1

    ws1.merge_cells('A'+str(baris)+':H'+str(baris))
    for col in ws1['A'+str(baris):'H'+str(baris)]:
        for cell in col:
            cell.style = header3
    baris +=1    

    ws1.merge_cells('A'+str(baris)+':F'+str(baris))
    for col in ws1['A'+str(baris):'H'+str(baris)]:
        for cell in col:
            cell.style = header1
    ws1.row_dimensions[baris].height = 20
    ws1['A'+str(baris)] = 'JUMLAH'
    ws1['G'+str(baris)] = "=SUM(G13:G"+str(baris-2)+")"

    baris +=2

    ws1.row_dimensions[baris].height = 20
    ws1['B'+str(baris)] = 'Kesepakatan Target'
    ws1.merge_cells('B'+str(baris)+':C'+str(baris))
    ws1['B'+str(baris)].style = boldkiri

    baris +=1

    ws1.row_dimensions[baris].height = 20
    ws1.merge_cells('B'+str(baris)+':C'+str(baris))
    ws1['B'+str(baris)].style = default_style
    ws1['B'+str(baris)] = 'Tanggal: 1 ' + dokumen_ckp.periode.strftime("%B %Y")

    baris +=2

    ws1.merge_cells('B'+str(baris)+':C'+str(baris))
    ws1.merge_cells('E'+str(baris)+':H'+str(baris))
    ws1.row_dimensions[baris].height = 20
    ws1['B'+str(baris)].style = tengah
    ws1['E'+str(baris)].style = tengah
    ws1['B'+str(baris)] = 'Pegawai Yang Dinilai'
    ws1['E'+str(baris)] = 'Pejabat Penilai'

    baris +=4

    ws1.merge_cells('B'+str(baris)+':C'+str(baris))
    ws1.merge_cells('E'+str(baris)+':H'+str(baris))
    ws1.row_dimensions[baris].height = 20
    ws1['B'+str(baris)].style = boldtengah
    ws1['E'+str(baris)].style = boldtengah
    ws1['B'+str(baris)] = dokumen_ckp.pegawai.get_full_name()
    ws1['E'+str(baris)] = 'Ir. Budianto'

    baris +=1

    ws1.merge_cells('B'+str(baris)+':C'+str(baris))
    ws1.merge_cells('E'+str(baris)+':H'+str(baris))
    ws1.row_dimensions[baris].height = 20
    ws1['B'+str(baris)].style = boldtengah
    ws1['E'+str(baris)].style = boldtengah
    ws1['B'+str(baris)] = 'NIP: ' +  dokumen_ckp.pegawai.nip
    ws1['E'+str(baris)] = 'NIP: ' + '196707261994011001'

    for row in range(baris-10, baris):
        ws1.row_dimensions[row].height = 20

    # Sheet CKP-T ----------------------------------------------------------------------------------------------------------------- #

    # Sheet CKP-R -----------------------------------------------------------------------------------------------------------------

    ws2 = wb.create_sheet("CKP-R")

    ws2.sheet_format.defaultRowHeight = 20
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 17
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 12
    ws2.column_dimensions['J'].width = 12
    ws2.column_dimensions['K'].width = 50

     # Merge Cell
    ws2.merge_cells('A2:K2')
    ws2.merge_cells('A4:B4')
    ws2.merge_cells('A5:B5')
    ws2.merge_cells('A6:B6')
    ws2.merge_cells('A7:B7')
    ws2.merge_cells('B9:C10')
    ws2.merge_cells('A9:A10')
    ws2.merge_cells('D9:D10')
    ws2.merge_cells('E9:G9')
    ws2.merge_cells('H9:H10')
    ws2.merge_cells('I9:I10')
    ws2.merge_cells('J9:J10')
    ws2.merge_cells('K9:K10')
    ws2.merge_cells('B11:C11')
    ws2.merge_cells('A12:K12')

    for col in ws2['A4':'C7']:
        for cell in col:
            cell.style = default_style

    ws2.row_dimensions[1].height = 30
    ws2['K1'].style = tengahboldborder
    ws2.row_dimensions[2].height = 25
    ws2['A2'].style = tengahbold
    ws2.row_dimensions[9].height = 25
    ws2.row_dimensions[10].height = 25
    for col in ws2['A9':'K10']:
        for cell in col:
            cell.style = header1
    ws2.row_dimensions[11].height = 20
    for col in ws2['A11':'K11']:
        for cell in col:
            cell.style = header2
    ws2.row_dimensions[12].height = 20
    for col in ws2['A12':'K12']:
        for cell in col:
            cell.style = header3
    for i in range(4,9):
        ws2.row_dimensions[i].height = 17

    ws2['K1'] = "CKP-R"
    ws2['A2'] = "CAPAIAN KINERJA PEGAWAI TAHUN " + str(dokumen_ckp.periode.year)
    ws2['A4'] = "Satuan Organisasi"
    ws2['C4'] = ": BPS Kabupaten Kuantan Singingi"
    ws2['A5'] = "Nama"
    ws2['C5'] = ": " + dokumen_ckp.pegawai.get_full_name()
    ws2['A6'] = "Jabatan"
    ws2['C6'] = ": " + dokumen_ckp.pegawai.get_jabatan_kantor_display()
    ws2['A7'] = "Periode"
    ws2['C7'] = ": 1 " + dokumen_ckp.periode.strftime("%B ") + "- " + str(last_day_of_month) + " " +dokumen_ckp.periode.strftime("%B %Y")

    ws2['A9'] = "No"
    ws2['B9'] = "Uraian Kegiatan"
    ws2['D9'] = "Satuan"
    ws2['E9'] = "Kuantitas" 
    ws2['E10'] = "Target" 
    ws2['F10'] = "Realisasi" 
    ws2['G10'] = "%" 
    ws2['H9'] = "Tingkat Kualitas (%)"
    ws2['I9'] = "Kode Butir Kegiatan"
    ws2['J9'] = "Angka Kredit"
    ws2['K9'] = "Keterangan"
    ws2['A11'] = "(1)"
    ws2['B11'] = "(2)"
    ws2['D11'] = "(3)"
    ws2['E11'] = "(4)"
    ws2['F11'] = "(5)"
    ws2['G11'] = "(6)"
    ws2['H11'] = "(7)"
    ws2['I11'] = "(8)"
    ws2['J11'] = "(9)"
    ws2['K11'] = "(10)"
    ws2['A12'] = "UTAMA"

    baris = 13

    nomor = 1

    for idx, butir_ckp in enumerate(butir_ckp_set):

        if butir_ckp.jenis_butir_ckp == 'TAMBAHAN':
            continue

        ws2.merge_cells('B'+str(baris)+':C'+str(baris))

        ws2['A'+str(baris)] = nomor
        ws2['B'+str(baris)] = butir_ckp.kegiatan.uraian_kegiatan
        ws2['D'+str(baris)] = butir_ckp.kegiatan.satuan
        ws2['E'+str(baris)] = butir_ckp.target
        ws2['F'+str(baris)] = butir_ckp.realisasi
        if butir_ckp.realisasi:
            ws2['G'+str(baris)] = "=F"+str(baris)+"/E"+str(baris)+"*100" 
        ws2['H'+str(baris)] = butir_ckp.tingkat_kualitas
        if butir_ckp.kegiatan.butir_kegiatan is not None:
            ws2['I'+str(baris)] = butir_ckp.kegiatan.butir_kegiatan.__str__()
            ws2['J'+str(baris)] = "=" + str(butir_ckp.kegiatan.butir_kegiatan.angka_kredit) + "*F"+str(baris)
        ws2['K'+str(baris)] = butir_ckp.keterangan

        for col in ws2['A'+str(baris):'K'+str(baris)]:
            for cell in col:
                cell.style = isitabeltengah

        ws2['B'+str(baris)].style = isitabelkiri
        ws2['E'+str(baris)].style = isitabelkanan
        ws2['F'+str(baris)].style = isitabelkanan
        ws2['K'+str(baris)].style = isitabelkiri
        
        ws2.row_dimensions[baris].height = 25
        if len(butir_ckp.kegiatan.uraian_kegiatan)>45 :
            ws2.row_dimensions[baris].height = 50
        
        ws2['G'+str(baris)].number_format = '0.00'
        ws2['H'+str(baris)].number_format = '0.00'
        ws2['J'+str(baris)].number_format = '0.000'

        baris +=1
        nomor +=1

    ws2.merge_cells('A'+str(baris)+':K'+str(baris))
    for col in ws2['A'+str(baris):'K'+str(baris)]:
        for cell in col:
            cell.style = header3
    ws2.row_dimensions[baris].height = 20
    baris +=1
    ws2.row_dimensions[baris].height = 20
    for col in ws2['A'+str(baris):'K'+str(baris)]:
        for cell in col:
            cell.style = header3
    ws2['A'+str(baris)] = 'TAMBAHAN'
    ws2.merge_cells('A'+str(baris)+':K'+str(baris))  
    baris +=1
    
    nomor = 1

    for idx, butir_ckp in enumerate(butir_ckp_set):

        if butir_ckp.jenis_butir_ckp == 'UTAMA':
            continue

        ws2.merge_cells('B'+str(baris)+':C'+str(baris))
        
        ws2['A'+str(baris)] = nomor
        ws2['B'+str(baris)] = butir_ckp.kegiatan.uraian_kegiatan
        ws2['D'+str(baris)] = butir_ckp.kegiatan.satuan
        ws2['E'+str(baris)] = butir_ckp.target
        ws2['F'+str(baris)] = butir_ckp.realisasi
        if butir_ckp.realisasi:
            ws2['G'+str(baris)] = "=F"+str(baris)+"/E"+str(baris)+"*100"
        ws2['H'+str(baris)] = butir_ckp.tingkat_kualitas
        if butir_ckp.kegiatan.butir_kegiatan is not None:
            ws2['I'+str(baris)] = butir_ckp.kegiatan.butir_kegiatan.__str__()
            ws2['J'+str(baris)] = "=" + str(butir_ckp.kegiatan.butir_kegiatan.angka_kredit) + "*F"+str(baris)
        ws2['K'+str(baris)] = butir_ckp.keterangan

        for col in ws2['A'+str(baris):'K'+str(baris)]:
            for cell in col:
                cell.style = isitabeltengah

        ws2['B'+str(baris)].style = isitabelkiri
        ws2['E'+str(baris)].style = isitabelkanan
        ws2['F'+str(baris)].style = isitabelkanan
        ws2['K'+str(baris)].style = isitabelkiri
        
        ws2.row_dimensions[baris].height = 25
        if len(butir_ckp.kegiatan.uraian_kegiatan)>45 :
            ws2.row_dimensions[baris].height = 50
        
        ws2['G'+str(baris)].number_format = '0.00'
        ws2['H'+str(baris)].number_format = '0.00'
        ws2['J'+str(baris)].number_format = '0.000'


        baris +=1
        nomor +=1

    ws2.merge_cells('A'+str(baris)+':K'+str(baris))
    for col in ws2['A'+str(baris):'K'+str(baris)]:
        for cell in col:
            cell.style = header3
    baris +=1    

    ws2.merge_cells('A'+str(baris)+':I'+str(baris))
    for col in ws2['A'+str(baris):'K'+str(baris+2)]:
        for cell in col:
            cell.style = header1
    ws2.row_dimensions[baris].height = 20
    ws2['A'+str(baris)] = 'JUMLAH'
    ws2['J'+str(baris)] = "=SUM(J13:J"+str(baris-2)+")"
    ws2['J'+str(baris)].number_format = '0.000'

    baris +=1

    ws2.merge_cells('A'+str(baris)+':F'+str(baris))
    ws2.row_dimensions[baris].height = 20
    ws2['A'+str(baris)] = 'RATA-RATA'
    ws2['G'+str(baris)] = "=AVERAGE(G13:G"+str(baris-3)+")"
    ws2['H'+str(baris)] = "=AVERAGE(H13:H"+str(baris-3)+")"
    ws2['G'+str(baris)].number_format = '0.00'
    ws2['H'+str(baris)].number_format = '0.00'

    baris +=1

    ws2.merge_cells('A'+str(baris)+':F'+str(baris))
    ws2.merge_cells('G'+str(baris)+':H'+str(baris))
    ws2.row_dimensions[baris].height = 20
    ws2['A'+str(baris)] = 'CAPAIAN KINERJA PEGAWAI (CKP)'
    ws2['G'+str(baris)] = "=AVERAGE(G"+str(baris-1)+":H"+str(baris-1)+")"
    ws2['G'+str(baris)].number_format = '0.00'

    ws2.merge_cells('I'+str(baris-1)+':I'+str(baris))
    ws2.merge_cells('J'+str(baris-1)+':J'+str(baris))
    ws2.merge_cells('K'+str(baris-1)+':K'+str(baris))

    baris +=2

    ws2.row_dimensions[baris].height = 20
    ws2['B'+str(baris)] = 'Penilaian Kinerja'
    ws2.merge_cells('B'+str(baris)+':C'+str(baris))
    ws2['B'+str(baris)].style = boldkiri

    baris +=1

    ws2.row_dimensions[baris].height = 20
    ws2.merge_cells('B'+str(baris)+':C'+str(baris))
    ws2['B'+str(baris)].style = default_style
    ws2['B'+str(baris)] = 'Tanggal: ' + str(last_day_of_month) + ' ' + dokumen_ckp.periode.strftime("%B %Y")

    baris = baris+2
    
    ws2.merge_cells('B'+str(baris)+':C'+str(baris))
    ws2.merge_cells('G'+str(baris)+':K'+str(baris))
    ws2.row_dimensions[baris].height = 20
    ws2['B'+str(baris)].style = tengah
    ws2['G'+str(baris)].style = tengah
    ws2['B'+str(baris)] = 'Pegawai Yang Dinilai'
    ws2['G'+str(baris)] = 'Pejabat Penilai'

    baris = baris+4

    ws2.merge_cells('B'+str(baris)+':C'+str(baris))
    ws2.merge_cells('G'+str(baris)+':K'+str(baris))
    ws2.row_dimensions[baris].height = 20
    ws2['B'+str(baris)].style = boldtengah
    ws2['G'+str(baris)].style = boldtengah
    ws2['B'+str(baris)] = dokumen_ckp.pegawai.get_full_name()
    ws2['G'+str(baris)] = 'Ir. Budianto'

    baris +=1

    ws2.merge_cells('B'+str(baris)+':C'+str(baris))
    ws2.merge_cells('G'+str(baris)+':K'+str(baris))
    ws2.row_dimensions[baris].height = 20
    ws2['B'+str(baris)].style = boldtengah
    ws2['G'+str(baris)].style = boldtengah
    ws2['B'+str(baris)] = 'NIP: ' +  dokumen_ckp.pegawai.nip
    ws2['G'+str(baris)] = 'NIP: ' + '196707261994011001'

    for row in range(baris-10, baris):
        ws2.row_dimensions[row].height = 20
    
    # Sheet CKP-R ----------------------------------------------------------------------------------------------------------------- #

    wb.save(response)

    return response


# AJAX

def load_satuan(request):
    kegiatan = MasterKegiatan.objects.get(id=request.GET.get('kegiatan_id'))

    context = {
        'kegiatan': kegiatan,
    }

    return render(request, 'backend/ckp/satuan_field.html', context)


def load_butir_kegiatan(request):
    kegiatan = MasterKegiatan.objects.get(id=request.GET.get('kegiatan_id'))

    context = {
        'kegiatan': kegiatan,
    }

    return render(request, 'backend/ckp/butir_kegiatan_field.html', context)


def load_angka_kredit(request):
    kegiatan = MasterKegiatan.objects.get(id=request.GET.get('kegiatan_id'))

    context = {
        'kegiatan': kegiatan,
    }

    return render(request, 'backend/ckp/angka_kredit_field.html', context)

def load_master_angka_kredit(request):
    butir_kegiatan = MasterButirKegiatan.objects.get(id=request.GET.get('butir_kegiatan_id'))

    context = {
        'butir_kegiatan': butir_kegiatan,
    }

    return render(request, 'backend/ckp/master_angka_kredit_field.html', context)

# <------------------------------------------ END SICAKEP - CKP ----------------------------------------->

# <------------------------------------------ SIPIA --------------------------------------------->
@login_required
def daftar_penilaian_pia(request):
    pegawai_list = CustomUser.objects.all().exclude(is_superuser=True).exclude(id=request.user.id).extra(select={"jabatan_kantor_int": "CAST(jabatan_kantor AS INTEGER)"}).order_by("jabatan_kantor_int")

    context = {
        'pegawai_list': pegawai_list,
    }
    return render(request, 'backend/pia/daftarpenilaianpia.html', context)

# <------------------------------------------ END SIPIA --------------------------------------------->
# <app>/<model>_<viewtype>.html
